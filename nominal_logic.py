from __future__ import annotations

import csv
import hashlib
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DEFAULT_HOME_MAPPING: Dict[str, str] = {
    "2": "CRE",
    "3": "PET",
    "4": "ATW",
    "5": "BUN",
    "6": "ALI",
    "7": "HO",
}

HOME_NAME_TO_CODE: Dict[str, str] = {
    "crescent": "CRE",
    "crescent nursing home": "CRE",
    "peter": "PET",
    "peters": "PET",
    "peter's place": "PET",
    "peters place": "PET",
    "atwell": "ATW",
    "atwell house": "ATW",
    "bunyan": "BUN",
    "bunyan lodge": "BUN",
    "alicia": "ALI",
    "alicia nursing home": "ALI",
    "head office": "HO",
    "office": "HO",
}

PAYROLL_LINE_RE = re.compile(
    r"^(?P<nic>Employers\s+NIC\s+on\s+)?(?P<category>.+?)\s+for\s+(?P<name>.+?)\s+-\s+Period\s+(?P<period>\d+)\s*$",
    re.IGNORECASE,
)


@dataclass
class ParsedLine:
    row_no: int
    nominal_code: str
    reference: str
    debit: float
    credit: float
    amount: float
    description: str
    is_nic: bool
    category: str
    is_other: bool
    staff_name: str
    staff_key: str
    period: str
    source_file: str


def file_signature(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def clean_nominal(value: Any) -> str:
    text = clean_text(value)
    if not text or text.lower() == "nan":
        return ""
    match = re.search(r"\d+", text)
    return match.group(0) if match else ""


def validate_nominal(value: Any) -> Tuple[bool, str]:
    code = clean_nominal(value)
    if not code:
        return False, "Nominal code cannot be blank. Example: 208, 404, 504."
    if not re.fullmatch(r"\d{3,5}", code):
        return False, "Nominal code must be 3 to 5 digits only. Example: 208, 404, 504."
    return True, code


def clean_amount(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\u00a3", "").replace("\u7a16", "").replace(",", "")
    if not text or text.lower() == "nan":
        return 0.0
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1].strip()
    try:
        return float(text)
    except ValueError:
        return 0.0


def staff_key(name: str) -> str:
    text = re.sub(r"[^a-zA-Z ]", " ", name.lower())
    return re.sub(r"\s+", " ", text).strip()


def staff_token_key(name: str) -> str:
    tokens = [t for t in staff_key(name).split() if t]
    return " ".join(sorted(tokens))


def staff_lookup_keys(name: str) -> List[str]:
    base = staff_key(name)
    tokens = [t for t in base.split() if t]
    no_initials = [t for t in tokens if len(t) > 1]
    keys = {base, staff_token_key(name), " ".join(no_initials), " ".join(sorted(no_initials))}
    if len(no_initials) >= 2:
        keys.add(f"{no_initials[0]} {no_initials[1]}")
        keys.add(f"{no_initials[-1]} {no_initials[0]}")
        keys.add(f"{no_initials[0]} {no_initials[-1]}")
        for i in range(len(no_initials) - 1):
            a, b = no_initials[i], no_initials[i + 1]
            keys.add(f"{a} {b}")
            keys.add(f"{b} {a}")
    return [k for k in keys if k]


def normalise_home(value: Any) -> str:
    text = staff_key(clean_text(value))
    if not text:
        return ""
    for alias, code in HOME_NAME_TO_CODE.items():
        if staff_key(alias) == text:
            return code
    return ""


def nominal_home(nominal_code: str, mapping: Dict[str, str]) -> str:
    code = clean_nominal(nominal_code)
    if not code:
        return ""
    return str(mapping.get(code[0], "UNKNOWN"))


def amount_from_debit_credit(debit: float, credit: float) -> float:
    return debit - abs(credit)


def is_other_category(category: str) -> bool:
    return "OTHER" in category.upper()


def parse_description(
    row_no: int,
    nominal: str,
    reference: Any,
    debit: float,
    credit: float,
    desc_value: Any,
    source_file: str,
) -> Optional[ParsedLine]:
    desc = clean_text(desc_value)
    match = PAYROLL_LINE_RE.match(desc)
    if not match:
        return None
    category = clean_text(match.group("category")).upper()
    name = clean_text(match.group("name"))
    return ParsedLine(
        row_no=row_no,
        nominal_code=clean_nominal(nominal),
        reference=clean_text(reference),
        debit=debit,
        credit=credit,
        amount=amount_from_debit_credit(debit, credit),
        description=desc,
        is_nic=bool(match.group("nic")),
        category=category,
        is_other=is_other_category(category),
        staff_name=name,
        staff_key=staff_key(name),
        period=clean_text(match.group("period")),
        source_file=source_file,
    )


def find_header_row(rows: List[List[Any]]) -> int:
    for idx, row in enumerate(rows):
        values = [staff_key(clean_text(v)) for v in row]
        joined = " | ".join(values)
        if "code" in values and "description" in joined and ("debit" in values or "credit" in values):
            return idx
    raise ValueError("Could not find the header row. Expected columns like Code, Reference, Debit, Credit, Description.")


def detect_columns(row: List[Any]) -> Dict[str, int]:
    result: Dict[str, int] = {}
    for i, value in enumerate(row):
        key = staff_key(clean_text(value))
        if key == "code":
            result["code"] = i
        elif key == "reference":
            result["reference"] = i
        elif key == "debit":
            result["debit"] = i
        elif key == "credit":
            result["credit"] = i
        elif key == "description":
            result["description"] = i
    missing = [x for x in ["code", "description", "debit", "credit"] if x not in result]
    if missing:
        raise ValueError("Missing expected column(s): " + ", ".join(missing))
    return result


def read_csv_rows(path: Path) -> List[List[Any]]:
    last_error: Optional[Exception] = None
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with path.open(newline="", encoding=enc) as f:
                return list(csv.reader(f))
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error:
        raise last_error
    return []


def read_excel_rows(path: Path) -> List[List[Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


def read_lines(input_file: Path) -> List[ParsedLine]:
    rows = read_csv_rows(input_file) if input_file.suffix.lower() == ".csv" else read_excel_rows(input_file)
    header_idx = find_header_row(rows)
    cols = detect_columns(rows[header_idx])
    lines: List[ParsedLine] = []
    last_nominal = ""

    for zero_idx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 1):
        excel_row = zero_idx + 1

        def get(col_name: str) -> Any:
            col = cols.get(col_name)
            if col is None or col >= len(row):
                return ""
            return row[col]

        raw_nominal = clean_nominal(get("code"))
        if raw_nominal:
            last_nominal = raw_nominal
        parsed = parse_description(
            row_no=excel_row,
            nominal=raw_nominal or last_nominal,
            reference=get("reference"),
            debit=clean_amount(get("debit")),
            credit=clean_amount(get("credit")),
            desc_value=get("description"),
            source_file=input_file.name,
        )
        if parsed:
            lines.append(parsed)
    return lines


def detect_staff_master_columns(rows: List[List[Any]]) -> Tuple[int, Dict[str, int]]:
    aliases = {
        "employee_code": {"employee code", "emp code", "code", "employee no", "payroll no"},
        "department": {"department code", "department", "department code2", "home", "site"},
        "surname": {"surname", "last name"},
        "forename": {"forename 1", "forename", "first name"},
        "full_name": {"full name", "staff name", "employee name", "name"},
        "nominal": {"nominal", "nominal code", "base nominal"},
    }
    for row_idx, row in enumerate(rows[:25]):
        found: Dict[str, int] = {}
        for col_idx, value in enumerate(row):
            key = staff_key(clean_text(value))
            for target, names in aliases.items():
                if key in names and target not in found:
                    found[target] = col_idx
        if "employee_code" in found or "full_name" in found or ("surname" in found and "forename" in found):
            return row_idx, found
    raise ValueError("Could not detect columns in the staff master. Expected Employee Code, Department Code, Surname, Forename or Full Name.")


def load_staff_master(master_file: Optional[Path], mapping: Dict[str, str]) -> Dict[str, Dict[str, str]]:
    if not master_file or not master_file.exists():
        return {}
    wb = load_workbook(master_file, data_only=True, read_only=True)
    ws_name = "List Data" if "List Data" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[ws_name]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    header_idx, cols = detect_staff_master_columns(rows)
    result: Dict[str, Dict[str, str]] = {}

    for row in rows[header_idx + 1 :]:
        def get(col_name: str) -> str:
            col = cols.get(col_name)
            if col is None or col >= len(row):
                return ""
            return clean_text(row[col])

        full_name = get("full_name") or clean_text(f"{get('forename')} {get('surname')}")
        if not full_name:
            continue
        explicit_nominal = clean_nominal(get("nominal"))
        base_nominal = explicit_nominal if explicit_nominal else ""
        department = get("department")
        base_home = nominal_home(base_nominal, mapping) if base_nominal else normalise_home(department)
        master_source = "STAFF_MASTER_NOMINAL" if base_nominal else ("STAFF_MASTER_HOME" if base_home else "STAFF_MASTER")
        item = {
            "master_staff_name": full_name,
            "employee_code": clean_nominal(get("employee_code")),
            "base_nominal": base_nominal,
            "base_home": base_home,
            "department": department,
            "source": master_source,
        }
        for k in staff_lookup_keys(full_name):
            result[k] = item

        surname = get("surname")
        forename = get("forename")
        if surname or forename:
            first_forename = forename.split()[0] if forename else ""
            for candidate in [f"{forename} {surname}", f"{surname} {forename}", f"{first_forename} {surname}", f"{surname} {first_forename}"]:
                for k in staff_lookup_keys(candidate):
                    result[k] = item
    return result


def review_lines(lines: Iterable[ParsedLine]) -> List[ParsedLine]:
    return [line for line in lines if line.is_other]


def should_use_for_base(line: ParsedLine) -> bool:
    if line.is_nic:
        return False
    if line.is_other:
        return False
    return bool(line.nominal_code)


def infer_staff_base(
    lines: Iterable[ParsedLine],
    mapping: Dict[str, str],
    staff_master: Dict[str, Dict[str, str]],
    overrides: Dict[str, str],
) -> Dict[str, Dict[str, Any]]:
    totals: Dict[str, Dict[str, float]] = defaultdict(lambda: defaultdict(float))
    display: Dict[str, str] = {}
    for line in lines:
        key = line.staff_key
        if line.is_other:
            display.setdefault(key, line.staff_name)
        if should_use_for_base(line):
            totals[key][line.nominal_code] += abs(line.amount)

    result: Dict[str, Dict[str, Any]] = {}
    for key, name in display.items():
        override = clean_nominal(overrides.get(key, ""))
        master: Dict[str, str] = {}
        for lookup_key in staff_lookup_keys(name):
            found_master = staff_master.get(lookup_key)
            if found_master:
                master = found_master
                break

        source = "UNKNOWN"
        base_nominal = ""
        base_home = ""
        confidence = "LOW"
        total_used = 0.0

        if override:
            base_nominal = override
            base_home = nominal_home(base_nominal, mapping)
            source = "MANUAL_OVERRIDE"
            confidence = "MANUAL"
        elif master:
            base_nominal = clean_nominal(master.get("base_nominal", ""))
            base_home = clean_text(master.get("base_home", "")) or nominal_home(base_nominal, mapping)
            source = clean_text(master.get("source", "STAFF_MASTER"))
            confidence = "HIGH" if base_home else "LOW"
        elif totals.get(key):
            code_totals = totals[key]
            best_code = max(code_totals.keys(), key=lambda code: code_totals[code])
            base_nominal = best_code
            base_home = nominal_home(best_code, mapping)
            source = "INFERRED_FROM_REPORT"
            confidence = "MEDIUM"
            total_used = round(code_totals[best_code], 2)

        result[key] = {
            "staff_name": name,
            "base_nominal": base_nominal,
            "base_home": base_home,
            "source": source,
            "confidence": confidence,
            "amount_used_for_inference": total_used,
            "employee_code": clean_text(master.get("employee_code", "")) if master else "",
            "department": clean_text(master.get("department", "")) if master else "",
        }
    return result


def missing_nominal_key(line: ParsedLine) -> str:
    return f"{line.staff_key}|{line.category}|P{line.period}|{round(abs(line.amount), 2)}|{'NIC' if line.is_nic else 'WAGES'}"


def apply_missing_nominals(lines: List[ParsedLine], corrections: Dict[str, str]) -> None:
    for line in lines:
        if line.is_other and not line.nominal_code:
            saved = clean_nominal(corrections.get(missing_nominal_key(line), ""))
            if saved:
                line.nominal_code = saved


def nominal_dropdown_options(lines: List[ParsedLine], mapping: Dict[str, str]) -> List[str]:
    existing = {clean_nominal(line.nominal_code) for line in lines if clean_nominal(line.nominal_code)}
    suffixes = {code[1:] for code in existing if len(code) >= 3}
    for prefix in mapping.keys():
        if re.fullmatch(r"\d", str(prefix)):
            for suffix in suffixes:
                if suffix:
                    existing.add(f"{prefix}{suffix}")
    return sorted(existing, key=lambda x: (len(x), x))


def build_review_rows(
    lines: List[ParsedLine],
    base: Dict[str, Dict[str, Any]],
    mapping: Dict[str, str],
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    rows: List[Dict[str, Any]] = []
    for line in review_lines(lines):
        b = base.get(line.staff_key, {})
        base_nominal = clean_text(b.get("base_nominal", ""))
        base_home = clean_text(b.get("base_home", ""))
        other_home = nominal_home(line.nominal_code, mapping)
        cost_type = "NIC" if line.is_nic else "WAGES"
        status = "REVIEW_BLANK_BASE"
        if base_home and other_home:
            if other_home == "UNKNOWN":
                status = "REVIEW_UNKNOWN_NOMINAL_SERIES"
            elif base_home == other_home:
                status = "SAME_HOME_OTHER_JOB_OR_SAME_HOME_OTHER"
            else:
                status = "TRANSFER_TO_OTHER_HOME"

        rows.append({
            "Status": status,
            "Staff Name": line.staff_name,
            "Period": line.period,
            "Cost Type": cost_type,
            "Other Category": line.category,
            "Payroll Nominal": line.nominal_code,
            "Payroll Nominal Home": other_home,
            "Base Nominal": base_nominal,
            "Base Home": base_home,
            "Transfer From": base_home if status == "TRANSFER_TO_OTHER_HOME" else "",
            "Transfer To": other_home if status == "TRANSFER_TO_OTHER_HOME" else "",
            "Amount": round(abs(line.amount), 2),
            "Debit": line.debit,
            "Credit": line.credit,
            "Description": line.description,
            "Reference": line.reference,
            "Row No": line.row_no,
            "Base Source": b.get("source", ""),
            "Confidence": b.get("confidence", ""),
        })

    review_df = pd.DataFrame(rows)
    transfer_df = pd.DataFrame()
    if not review_df.empty:
        transfer_df = (
            review_df[review_df["Status"] == "TRANSFER_TO_OTHER_HOME"]
            .groupby(["Period", "Transfer From", "Transfer To", "Cost Type"], dropna=False)["Amount"]
            .sum()
            .reset_index()
            .sort_values(["Period", "Transfer From", "Transfer To", "Cost Type"])
        )
    nominal_df = pd.DataFrame()
    if not review_df.empty:
        nominal_df = (
            review_df.groupby(["Payroll Nominal", "Payroll Nominal Home", "Cost Type"], dropna=False)["Amount"]
            .sum()
            .reset_index()
            .sort_values(["Payroll Nominal", "Cost Type"])
        )
    staff_df = pd.DataFrame(list(base.values())).sort_values("staff_name") if base else pd.DataFrame()
    return review_df, transfer_df, nominal_df, staff_df


def autosize_and_style(filename: Path) -> None:
    wb = load_workbook(filename)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    transfer_fill = PatternFill("solid", fgColor="E2F0D9")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    error_fill = PatternFill("solid", fgColor="F4CCCC")
    border = Border(bottom=Side(style="thin", color="D9E2F3"))

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        status_col: Optional[int] = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Status":
                status_col = idx
                break
        if status_col is not None:
            for row_idx in range(2, ws.max_row + 1):
                status_value = clean_text(ws.cell(row=row_idx, column=status_col).value)
                fill = None
                if status_value == "TRANSFER_TO_OTHER_HOME":
                    fill = transfer_fill
                elif status_value.startswith("REVIEW"):
                    fill = error_fill
                elif status_value.startswith("SAME_HOME"):
                    fill = warning_fill
                if fill is not None:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill
        for col_idx in range(1, ws.max_column + 1):
            max_len = 10
            for row_idx in range(1, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                max_len = max(max_len, len(clean_text(value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)
    wb.save(filename)


def write_output(
    output_file: Path,
    review_df: pd.DataFrame,
    transfer_df: pd.DataFrame,
    nominal_df: pd.DataFrame,
    staff_df: pd.DataFrame,
    mapping: Dict[str, str],
    corrections: Dict[str, str],
    overrides: Dict[str, str],
) -> Path:
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        review_df.to_excel(writer, sheet_name="OTHER_Review", index=False)
        transfer_df.to_excel(writer, sheet_name="Transfer_Summary", index=False)
        nominal_df.to_excel(writer, sheet_name="Nominal_Summary", index=False)
        staff_df.to_excel(writer, sheet_name="Staff_Base", index=False)
        pd.DataFrame([{"Nominal Prefix": k, "Home": v} for k, v in sorted(mapping.items())]).to_excel(writer, sheet_name="Nominal_Mapping", index=False)
        pd.DataFrame([{"This File Saved Key": k, "Nominal": v} for k, v in sorted(corrections.items())]).to_excel(writer, sheet_name="Missing_Nominal_This_File", index=False)
        pd.DataFrame([{"Staff Key": k, "Base Nominal": v} for k, v in sorted(overrides.items())]).to_excel(writer, sheet_name="Staff_Overrides", index=False)
    autosize_and_style(output_file)
    return output_file


def default_output_name(payroll_name: str) -> str:
    stem = Path(payroll_name).stem
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    return f"{stem}_OTHER_REVIEW_{stamp}.xlsx"
